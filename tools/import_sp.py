#!/usr/bin/env python3
"""import_sp.py — prevod zdrojových zošitov Sociálnej poisťovne na tidy CSV.

Zdroje sú v data/zdroj/ a sú commitnuté spolu s výstupom, takže sa dá kedykoľvek
overiť, čo z čoho vzniklo. Spusti po výmene ktoréhokoľvek zošita:

    python3 tools/import_sp.py

Čo tento skript rieši (a čo sa pri ručnom prepise ľahko pokazí):

1. POČTY SÚ V TISÍCACH. V mesačnom zošite je "Počet vyplatených SD" = 1134,69,
   čo je 1 134 690 dôchodkov. Overené proti ročnému zošitu, ktorý ten istý údaj
   uvádza v jednotkách. Prepočítavame na jednotky.

2. VÝDAVKY V ROČNÝCH ZOŠITOCH SÚ KUMULATÍVNE. Stĺpce sú "január", "január a
   február", …, "január až december". Ročná hodnota je teda posledný stĺpec, nie
   súčet stĺpcov.

3. DVE RÔZNE DEFINÍCIE PRIEMERNÉHO DÔCHODKU. Mesačný zošit má "Priemerný SD*",
   čo je aritmetický priemer sólo A kráteného dôchodku (dec. 2024: 642,35 €).
   Ročný zošit má "priemernú výšku vyplácaných SÓLO dôchodkov" (dec. 2024:
   683,10 €). Sú to iné veličiny a nesmú byť v jednom grafe — preto idú do dvoch
   samostatných CSV s odlišnou jednotkou v manifeste.

4. DEFINIČNÉ ZLOMY V RADE VÝDAVKOV:
     2021, 2022 — rodičovský dôchodok ešte neexistoval
     2023       — rodičovský vykázaný ako samostatný riadok
     2024       — "starobný dôchodok (vr. rodičovského)" ho už OBSAHUJE
   Pre porovnateľnú radu preto od roku 2024 rodičovský odčítavame a vykazujeme ho
   ako vlastnú kategóriu. Kontrolný súčet proti riadku "Celkom" to overuje.

5. POČET DÔCHODKOV ≠ POČET DÔCHODCOV. Jeden človek môže poberať viac dôchodkov
   (napr. starobný + vdovský), preto je "počet dôchodcov" nižší než "počet
   vyplácaných dôchodkov". Sú to dva samostatné datasety.
"""

from __future__ import annotations

import csv
import sys
from datetime import datetime
from pathlib import Path

try:
    import openpyxl
except ImportError:
    sys.exit("Chýba openpyxl. Nainštaluj: pip install openpyxl")

ROOT = Path(__file__).resolve().parent.parent
SRC = ROOT / "data" / "zdroj"
OUT = ROOT / "data"

MONTHLY_XLSX = SRC / "SP_socialne_davky_mesacne.xlsx"
YEARS = (2021, 2022, 2023, 2024)

MONTH_NAMES = {
    "január": 1, "február": 2, "marec": 3, "apríl": 4, "máj": 5, "jún": 6,
    "júl": 7, "august": 8, "september": 9, "október": 10, "november": 11,
    "december": 12,
}

# Poradie druhov dôchodku, ktoré sa drží vo všetkých výstupoch. Určuje aj
# priradenie farieb v grafoch (slot 1 = najväčšia položka).
DRUH_ORDER = [
    "Starobný", "Predčasný starobný", "Invalidný", "Vdovský", "Vdovecký",
    "Sirotský", "Rodičovský", "13. dôchodok",
]


def write_csv(name: str, header: list[str], rows: list[list], note: str = "") -> None:
    path = OUT / name
    with path.open("w", encoding="utf-8", newline="") as fh:
        w = csv.writer(fh, lineterminator="\n")
        w.writerow(header)
        w.writerows(rows)
    extra = f"  ({note})" if note else ""
    print(f"  {name:38s} {len(rows):5d} riadkov{extra}")


def r2(v):
    """Zaokrúhlenie na 2 desatinné miesta, celé čísla bez desatinnej časti."""
    if v is None:
        return None
    f = round(float(v), 2)
    return int(f) if f == int(f) else f


def load(path: Path):
    return openpyxl.load_workbook(path, data_only=True)


def month_iso(dt) -> str:
    return f"{dt.year:04d}-{dt.month:02d}"


# ─── 1. mesačný zošit: výdavky, počty, priemery, novopriznané ────────────────

# sheet -> (druh, výdavky, počet, priemer, novopriznané, priemer novopriznaný)
# Vdovský a vdovecký sú v jednom liste (VD a VDm), preto sa spracúva zvlášť.
MONTHLY_MAP = [
    ("cash mesacny SD", "Starobný",
     "Spolu výdavky na SD", "Počet vyplatených SD",
     "Priemerný SD*", "Počet novopriznaných SD (01-09P)",
     "Priemerný SD novopriznaný (01-03P)"),
    ("cash mesacny PSD", "Predčasný starobný",
     "Spolu výdavky na PSD", "Počet vyplatených PSD",
     "Priemerný PSD*", "Počet novopriznaných PSD (01-09P)",
     "Priemerný PSD novopriznaný (01-03P)"),
    ("cash mesacny SIR a zuctovanie", "Sirotský",
     "Spolu výdavky na SIR", None, None, None, None),
]


def sheet_index(ws) -> tuple[dict[str, int], list]:
    rows = list(ws.iter_rows(values_only=True))
    hdr = rows[1]
    idx = {}
    for j, h in enumerate(hdr):
        if h is not None:
            idx[str(h).strip()] = j
    return idx, rows[2:]


def cell(row, idx: dict, key: str | None):
    if key is None:
        return None
    j = idx.get(key)
    if j is None or j >= len(row):
        return None
    v = row[j]
    return v if isinstance(v, (int, float)) else None


def import_monthly(wb):
    vydavky, pocty, priemery, novo, priemery_novo = [], [], [], [], []

    def push(sheet, druh, k_vyd, k_poc, k_pri, k_novo, k_pri_novo):
        idx, body = sheet_index(wb[sheet])
        for row in body:
            if not isinstance(row[0], datetime):
                continue
            m = month_iso(row[0])
            v = cell(row, idx, k_vyd)
            if v is not None:
                vydavky.append([m, druh, r2(v)])
            p = cell(row, idx, k_poc)
            if p is not None:
                # tisíce -> jednotky
                pocty.append([m, druh, int(round(p * 1000))])
            a = cell(row, idx, k_pri)
            if a is not None:
                priemery.append([m, druh, r2(a)])
            n = cell(row, idx, k_novo)
            if n is not None:
                novo.append([m, druh, int(round(n * 1000))])
            an = cell(row, idx, k_pri_novo)
            if an is not None:
                priemery_novo.append([m, druh, r2(an)])

    for args in MONTHLY_MAP:
        push(*args)

    # invalidný: dva stupne (do 70 % a nad 70 %) — počty a priemery sú zvlášť,
    # výdavky len súhrnne, preto sa počty sčítavajú a priemer sa neuvádza za celok
    idx, body = sheet_index(wb["cash mesacny INV"])
    for row in body:
        if not isinstance(row[0], datetime):
            continue
        m = month_iso(row[0])
        v = cell(row, idx, "Spolu výdavky na INV")
        if v is not None:
            vydavky.append([m, "Invalidný", r2(v)])
        lo = cell(row, idx, "Počet vyplatených INV do 70 %")
        hi = cell(row, idx, "Počet vyplatených INV nad 70 %")
        if lo is not None and hi is not None:
            pocty.append([m, "Invalidný", int(round((lo + hi) * 1000))])
        nlo = cell(row, idx, "Počet novopriznaných INV do 70 % (01-09P)")
        nhi = cell(row, idx, "Počet novopriznaných INV nad 70 % (01-09P)")
        if nlo is not None and nhi is not None:
            novo.append([m, "Invalidný", int(round((nlo + nhi) * 1000))])
        # priemer za invalidný celok by bol vážený priemer dvoch stupňov;
        # vážime počtami, ktoré máme v tom istom riadku
        alo = cell(row, idx, "Priemerný INV do 70 %*")
        ahi = cell(row, idx, "Priemerný INV nad 70 %*")
        if None not in (alo, ahi, lo, hi) and (lo + hi) > 0:
            priemery.append([m, "Invalidný", r2((alo * lo + ahi * hi) / (lo + hi))])
        anlo = cell(row, idx, "Priemerný INV do 70 % novopriznaný (01-03P)")
        anhi = cell(row, idx, "Priemerný INV nad 70 % novopriznaný (01-03P)")
        if None not in (anlo, anhi, nlo, nhi) and (nlo + nhi) > 0:
            priemery_novo.append([m, "Invalidný", r2((anlo * nlo + anhi * nhi) / (nlo + nhi))])

    # vdovský + vdovecký z jedného listu
    idx, body = sheet_index(wb["cash mesacny VD"])
    for row in body:
        if not isinstance(row[0], datetime):
            continue
        m = month_iso(row[0])
        for druh, k_v, k_p, k_a, k_n, k_an in (
            ("Vdovský", "Spolu výdavky na VD", "Počet vyplatených VD", "Priemerný VD*",
             "Počet novopriznaných VD (01-09P)", "Priemerný VD novopriznaný (01-03P)"),
            ("Vdovecký", "Spolu výdavky na VDm", "Počet vyplatených VDm", "Priemerný VDm*",
             "Počet novopriznaných VDm (01-09P)", "Priemerný VDm novopriznaný (01-03P)"),
        ):
            v = cell(row, idx, k_v)
            if v is not None:
                vydavky.append([m, druh, r2(v)])
            p = cell(row, idx, k_p)
            if p is not None:
                pocty.append([m, druh, int(round(p * 1000))])
            a = cell(row, idx, k_a)
            if a is not None:
                priemery.append([m, druh, r2(a)])
            n = cell(row, idx, k_n)
            if n is not None:
                novo.append([m, druh, int(round(n * 1000))])
            an = cell(row, idx, k_an)
            if an is not None:
                priemery_novo.append([m, druh, r2(an)])

    # Rodičovský dôchodok má v mesačnom zošite vlastný stĺpec a "Spolu výdavky na
    # SD" ho NEOBSAHUJE. Bez tohto riadku chýbalo v roku 2023 presne 286 068 tis. €
    # proti ročnému zošitu.
    idx, body = sheet_index(wb["cash mesacny SD"])
    for row in body:
        if not isinstance(row[0], datetime):
            continue
        v = cell(row, idx, "rodičovský dôchodok")
        if v:
            vydavky.append([month_iso(row[0]), "Rodičovský", r2(v)])

    # 13. dôchodok je v mesačnom zošite rozdelený medzi druhy (ZFSP + ZFIP);
    # ako samostatná kategória výdavkov je zaujímavejší v súhrne
    trinasty: dict[str, float] = {}
    for sheet, key in (("cash mesacny SD", "13. dôchodok ZFSP"),
                       ("cash mesacny INV", "13. dôchodok ZFIP")):
        idx, body = sheet_index(wb[sheet])
        for row in body:
            if not isinstance(row[0], datetime):
                continue
            v = cell(row, idx, key)
            if v:
                m = month_iso(row[0])
                trinasty[m] = trinasty.get(m, 0) + v
    for m, v in sorted(trinasty.items()):
        vydavky.append([m, "13. dôchodok", r2(v)])

    order = {d: i for i, d in enumerate(DRUH_ORDER)}
    key = lambda r: (r[0], order.get(r[1], 99))
    return (sorted(vydavky, key=key), sorted(pocty, key=key),
            sorted(priemery, key=key), sorted(novo, key=key),
            sorted(priemery_novo, key=key))


# ─── 2. ročné zošity: výdavky, sólo priemery, počet dôchodcov ───────────────

# Riadky výdavkov, ako sú pomenované v jednotlivých rokoch -> jednotná kategória.
VYD_LABELS = {
    "starobný dôchodok": "Starobný",
    "starobný dôchodok (vr. rodičovského)": "Starobný",
    "predčasný starobný dôchodok": "Predčasný starobný",
    "invalidný dôchodok": "Invalidný",
    "vdovský dôchodok": "Vdovský",
    "vdovecký dôchodok": "Vdovecký",
    "sirotský dôchodok": "Sirotský",
    "rodičovský dôchodok": "Rodičovský",
    "z toho rodičovský dôchodok": "__rodicovsky_z_toho",
    "13. dôchodok": "13. dôchodok",
    "celkom": "__celkom",
}


def import_annual_expenditure() -> list[list]:
    out = []
    for year in YEARS:
        wb = load(SRC / f"SP_dochodkove_poistenie_{year}.xlsx")
        ws = next(w for w in wb.worksheets if "ýdavky" in w.title)
        vals, celkom, rodic_z_toho = {}, None, None

        for row in ws.iter_rows(values_only=True):
            label = row[1] if len(row) > 1 else None
            if not isinstance(label, str):
                continue
            key = VYD_LABELS.get(label.strip().lower())
            if key is None:
                continue
            # posledný stĺpec s číslom = kumulatív za január až december
            nums = [v for v in row[2:] if isinstance(v, (int, float))]
            if not nums:
                continue
            v = nums[-1]
            if key == "__celkom":
                celkom = v
            elif key == "__rodicovsky_z_toho":
                rodic_z_toho = v
            else:
                vals[key] = vals.get(key, 0) + v

        # 2024: rodičovský je súčasťou starobného -> vyčleniť, aby rada bola
        # porovnateľná s rokmi, kde bol samostatný alebo neexistoval
        if rodic_z_toho:
            vals["Starobný"] -= rodic_z_toho
            vals["Rodičovský"] = vals.get("Rodičovský", 0) + rodic_z_toho

        total = sum(vals.values())
        if celkom is not None and abs(total - celkom) > 1:
            sys.exit(f"KONTROLA ZLYHALA {year}: súčet druhov {total} != Celkom {celkom}")
        print(f"     {year}: kontrolný súčet OK ({total:,} tis. €)".replace(",", " "))

        for druh in DRUH_ORDER:
            if druh in vals:
                out.append([year, druh, r2(vals[druh])])
    return out


def import_annual_sheet(sheet_match: str, label_map: dict[str, str],
                        scale: float = 1.0, as_int: bool = False) -> list[list]:
    """Mesačné riadky z ročných zošitov (rovnaký layout vo všetkých rokoch)."""
    out = []
    for year in YEARS:
        wb = load(SRC / f"SP_dochodkove_poistenie_{year}.xlsx")
        ws = next((w for w in wb.worksheets if sheet_match in w.title), None)
        if ws is None:
            continue
        rows = list(ws.iter_rows(values_only=True))
        # hlavička je dvojriadková: r3 hlavné názvy, r4 podnázvy invalidného
        top = [str(c).strip() if c else "" for c in rows[3]]
        sub = [str(c).strip() if c else "" for c in rows[4]] if len(rows) > 4 else []

        # zlož efektívny názov stĺpca: "invalidný dôchodok" + "do 70%"
        names = []
        current = ""
        for j in range(len(top)):
            if top[j]:
                current = top[j]
            s = sub[j] if j < len(sub) else ""
            names.append(f"{current} {s}".strip() if s else current)

        for row in rows[5:]:
            if not isinstance(row[1], str):
                continue
            m = MONTH_NAMES.get(row[1].strip().lower())
            if m is None:
                continue
            month = f"{year:04d}-{m:02d}"
            for j, name in enumerate(names):
                key = label_map.get(name.lower())
                if key is None or j >= len(row):
                    continue
                v = row[j]
                if not isinstance(v, (int, float)):
                    continue
                val = v * scale
                out.append([month, key, int(round(val)) if as_int else r2(val)])
    order = {d: i for i, d in enumerate(DRUH_ORDER)}
    return sorted(out, key=lambda r: (r[0], order.get(r[1], 99)))


SOLO_LABELS = {
    "starobný dôchodok": "Starobný",
    "predčasný starobný dôchodok": "Predčasný starobný",
    "invalidný dôchodok spolu": "Invalidný",
    "invalidný dôchodok do 70%": "Invalidný do 70 %",
    "invalidný dôchodok nad 70%": "Invalidný nad 70 %",
    "vdovský dôchodok": "Vdovský",
    "vdovecký dôchodok": "Vdovecký",
    "sirotský dôchodok": "Sirotský",
}

DOCHODCOVIA_LABELS = {
    "starobný dôchodok": "Starobný",
    "predčasný starobný dôchodok": "Predčasný starobný",
    "invalidný dôchodok": "Invalidný",
    "vdovský dôchodok sólo": "Vdovský sólo",
    "vdovecký dôchodok sólo": "Vdovecký sólo",
    "sirotský dôchodok": "Sirotský",
    "spolu": "Spolu",
}


# ─── main ───────────────────────────────────────────────────────────────────

def main() -> None:
    if not MONTHLY_XLSX.exists():
        sys.exit(f"Chýba {MONTHLY_XLSX}")

    print("Prevádzam zdrojové zošity Sociálnej poisťovne:")

    wb = load(MONTHLY_XLSX)
    vydavky, pocty, priemery, novo, priemery_novo = import_monthly(wb)

    write_csv("sp_vydavky_mesacne.csv", ["mesiac", "druh", "vydavky_tis_eur"], vydavky)
    write_csv("sp_pocty_mesacne.csv", ["mesiac", "druh", "pocet"], pocty,
              "prepočítané z tisícov na jednotky")
    write_csv("sp_priemer_mesacne.csv", ["mesiac", "druh", "priemer_eur"], priemery,
              "aritm. priemer sólo a kráteného")
    write_csv("sp_novopriznane_mesacne.csv", ["mesiac", "druh", "pocet"], novo)

    # priemer vyplácaného vs novopriznaného v jednom dlhom CSV, aby sa dali
    # porovnať v jednom grafe na jednej osi (obe v EUR/mesiac)
    stav = {(m, d): v for m, d, v in priemery}
    novo_pr = {(m, d): v for m, d, v in priemery_novo}
    porovnanie = []
    for (m, d) in sorted(set(stav) & set(novo_pr), key=lambda k: (k[0], k[1])):
        porovnanie.append([m, d, "Vyplácané", stav[(m, d)]])
        porovnanie.append([m, d, "Novopriznané", novo_pr[(m, d)]])
    write_csv("sp_priemer_novo_vs_stav.csv",
              ["mesiac", "druh", "kategoria", "priemer_eur"], porovnanie)

    # celkové mesačné výdavky (súčet druhov) — pre sezónnosť a 3D povrch
    per_month: dict[str, float] = {}
    for m, _d, v in vydavky:
        per_month[m] = per_month.get(m, 0) + (v or 0)
    grid = [[int(m[:4]), int(m[5:7]), r2(v)] for m, v in sorted(per_month.items())]
    write_csv("sp_vydavky_rok_mesiac.csv", ["rok", "mesiac", "vydavky_tis_eur"], grid,
              "mriežka rok × mesiac pre teplotnú mapu a 3D povrch")

    print("   ročné zošity:")
    rocne = import_annual_expenditure()
    write_csv("sp_vydavky_rocne.csv", ["rok", "druh", "vydavky_tis_eur"], rocne,
              "z kumulatívneho stĺpca január–december")

    # waterfall: rozklad zmeny výdavkov medzi prvým a posledným ročným zošitom
    first, last = YEARS[0], YEARS[-1]
    by_year: dict[int, dict[str, float]] = {}
    for y, d, v in rocne:
        by_year.setdefault(y, {})[d] = v
    wf = [[f"Výdavky {first}", r2(sum(by_year[first].values())), 1]]
    for druh in DRUH_ORDER:
        delta = by_year[last].get(druh, 0) - by_year[first].get(druh, 0)
        if abs(delta) > 0.5:
            wf.append([druh, r2(delta), 0])
    wf.append([f"Výdavky {last}", r2(sum(by_year[last].values())), 1])
    write_csv("sp_vydavky_zmena.csv", ["faktor", "zmena_tis_eur", "je_uroven"], wf,
              f"rozklad zmeny {first} → {last}")

    solo = import_annual_sheet("priemerná výška", SOLO_LABELS)
    write_csv("sp_priemer_solo_mesacne.csv", ["mesiac", "druh", "priemer_eur"], solo,
              "iná definícia než sp_priemer_mesacne.csv — len sólo dôchodky")

    dochodcovia = import_annual_sheet("počet dôchodcov", DOCHODCOVIA_LABELS, as_int=True)
    write_csv("sp_dochodcovia_mesacne.csv", ["mesiac", "kategoria", "pocet"], dochodcovia,
              "osoby, nie dôchodky")

    # Dôchodky vs dôchodcovia v jednom CSV. Rozdiel medzi tými dvoma číslami je
    # počet ľudí poberajúcich viac než jeden dôchodok (typicky starobný +
    # vdovský), a je to najčastejšie zamieňaná dvojica v tejto štatistike.
    dochodky_spolu = import_annual_sheet("počet vyplácaných dôchodkov",
                                         {"spolu": "Vyplácané dôchodky"}, as_int=True)
    porovnaj = [r for r in dochodky_spolu]
    porovnaj += [[m, "Dôchodcovia (osoby)", p]
                 for m, kat, p in dochodcovia if kat == "Spolu"]
    porovnaj.sort(key=lambda r: (r[0], r[1]))
    write_csv("sp_dochodky_vs_dochodcovia.csv", ["mesiac", "ukazovatel", "pocet"], porovnaj,
              "jeden človek môže poberať viac dôchodkov")

    reconcile(vydavky, rocne)

    print("\nHotovo. Zdroj: data/zdroj/, výstup: data/sp_*.csv")


def reconcile(vydavky: list[list], rocne: list[list]) -> None:
    """Porovná súčet mesačných výdavkov s ročným zošitom po kategóriách.

    Dva zdroje nemusia sedieť na euro: mesačný zošit je hotovostná metodika,
    ročný je publikovaná ročná štatistika. Tento výpis však odhalí, keď niečo
    chýba systematicky — presne tak sa našlo, že "Spolu výdavky na SD"
    neobsahuje rodičovský dôchodok a v roku 2023 chýbalo 286 068 tis. €.
    """
    mon: dict[tuple[int, str], float] = {}
    for m, d, v in vydavky:
        key = (int(m[:4]), d)
        mon[key] = mon.get(key, 0) + (v or 0)
    ann: dict[tuple[int, str], float] = {}
    for y, d, v in rocne:
        ann[(y, d)] = ann.get((y, d), 0) + (v or 0)

    print("\n   kontrola mesačný (hotovostná metodika) vs ročný zošit:")
    tol = 0.005  # 0,5 % kategórie
    for year in YEARS:
        diffs = []
        for druh in DRUH_ORDER:
            m, a = mon.get((year, druh), 0), ann.get((year, druh), 0)
            if a == 0 and m == 0:
                continue
            d = m - a
            if abs(d) > max(1.0, abs(a) * tol):
                diffs.append(f"{druh} {d:+,.0f}".replace(",", " "))
        total_m = sum(v for (y, _), v in mon.items() if y == year)
        total_a = sum(v for (y, _), v in ann.items() if y == year)
        rel = (total_m - total_a) / total_a * 100 if total_a else 0
        status = "OK" if not diffs else "POZOR: " + ", ".join(diffs)
        print(f"     {year}: spolu {total_m:>12,.0f} vs {total_a:>12,.0f}"
              f" ({rel:+.2f} %)  {status}".replace(",", " "))


if __name__ == "__main__":
    main()
